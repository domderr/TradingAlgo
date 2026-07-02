from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_EXCEL = BASE_DIR / "Subscriptions.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "assets" / "subscriptions.json"

MARKET_ALIASES = {
    "italy": "Italy30",
    "italia": "Italy30",
    "all": "All",
}

INACTIVE_STATUSES = {"inactive", "disabled", "no", "false", "0", "off", "revoked"}


def normalize_market(value: str) -> str:
    token = value.strip()
    if not token:
        return ""
    return MARKET_ALIASES.get(token.lower(), token)


def parse_markets(value: object) -> list[str]:
    if value is None:
        return []
    raw = str(value).replace(";", ",")
    markets = [normalize_market(part) for part in raw.split(",")]
    return [market for market in markets if market]


def is_active(value: object) -> bool:
    if value is None or str(value).strip() == "":
        return True
    return str(value).strip().lower() not in INACTIVE_STATUSES


def password_hash(password: object) -> str:
    text = "" if password is None else str(password)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def build_subscriptions(excel_path: Path) -> list[dict[str, object]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    required = {"nome", "password", "markets"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    rows: list[dict[str, object]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        first_name = str(row.get("nome") or "").strip()
        password = row.get("password")
        markets = parse_markets(row.get("markets"))
        if not first_name and not password and not markets:
            continue
        if not first_name:
            raise ValueError("Found subscription row without nome")
        if password is None or str(password).strip() == "":
            raise ValueError(f"Missing password for {first_name}")
        if not markets:
            raise ValueError(f"Missing markets for {first_name}")
        rows.append(
            {
                "password_hash": password_hash(password),
                "first_name": first_name,
                "markets": markets,
                "active": is_active(row.get("status")),
            }
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate assets/subscriptions.json from Subscriptions.xlsx."
    )
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    subscriptions = build_subscriptions(args.excel)
    payload = {"subscriptions": subscriptions}
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(subscriptions)} subscriptions to {args.output}")


if __name__ == "__main__":
    main()
